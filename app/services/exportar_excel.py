from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from typing import List
from app.models.clase_programada import ClaseProgramada


def generar_excel_horario(clases: List[ClaseProgramada], nombre_docente: str) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Horario Docente"

    # Estilos
    encabezado_font = Font(bold=True, color="FFFFFF")
    encabezado_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    # Encabezados
    headers = ["Materia", "Aula", "Día", "Hora Inicio", "Hora Fin"]
    ws.append(headers)
    for col_num, col in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = encabezado_font
        cell.fill = encabezado_fill
        cell.alignment = center_align

    # Datos
    for clase in clases:
        ws.append([
            clase.materia.nombre,
            clase.aula,
            clase.dia,
            clase.hora_inicio.strftime("%H:%M"),
            clase.hora_fin.strftime("%H:%M"),
        ])

    # Ajuste de anchos
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
